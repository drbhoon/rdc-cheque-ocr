import os
import re
import json
import base64
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from functools import wraps
from datetime import date, datetime, timedelta
from flask import (Flask, request, jsonify, render_template, send_file,
                   session, redirect, url_for)
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
import anthropic

load_dotenv()

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024
app.secret_key = os.environ.get("SECRET_KEY", "dev-only-insecure-change-me")

ROLES = ("HO_ADMIN", "ACCOUNTS", "SALES", "VIEWER")
# Roles allowed to scan & upload cheques
UPLOAD_ROLES = ("HO_ADMIN", "ACCOUNTS", "SALES")


def landing_for(role):
    """Where a user lands after login / when sent off a forbidden page."""
    return "index" if role in UPLOAD_ROLES else "dashboard"

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "bmp", "pdf"}
STAFF_EXTENSIONS = {"xlsx", "xls"}

_client = None


# ── Anthropic client ────────────────────────────────────────────────────────

def get_client():
    global _client
    if _client is None:
        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            raise RuntimeError("ANTHROPIC_API_KEY is not set.")
        _client = anthropic.Anthropic(api_key=api_key)
    return _client


# ── Database helpers ─────────────────────────────────────────────────────────

def db_url():
    url = os.environ.get("DATABASE_URL", "")
    if url.startswith("postgres://"):
        url = "postgresql://" + url[len("postgres://"):]
    return url


def get_db():
    import psycopg2
    return psycopg2.connect(db_url())


def init_db():
    """Create / migrate all tables. Safe to run on every startup."""
    conn = get_db()
    try:
        cur = conn.cursor()

        # Base cheques table (Part I) — kept for backward compatibility
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cheques (
                id             SERIAL PRIMARY KEY,
                bank_name      TEXT,
                account_number TEXT,
                cheque_date    TEXT,
                payee          TEXT,
                amount_words   TEXT,
                amount_numbers TEXT,
                issuer_name    TEXT,
                scanned_at     TIMESTAMPTZ DEFAULT NOW()
            )
        """)

        # Part II additions — add columns if they don't exist yet
        cheque_cols = [
            ("cheque_number",     "TEXT"),
            ("cheque_date_iso",   "DATE"),
            ("deposit_due_date",  "DATE"),
            ("amount_value",      "NUMERIC"),
            ("status",            "TEXT DEFAULT 'PENDING'"),
            ("deposited_date",    "DATE"),
            ("deposit_bank",      "TEXT"),
            ("deposit_reference", "TEXT"),
            ("cleared_date",      "DATE"),
            ("bounce_date",       "DATE"),
            ("bounce_reason",     "TEXT"),
            ("sales_name",        "TEXT"),
            ("sales_email",       "TEXT"),
            ("location",          "TEXT"),
            ("plant",             "TEXT"),
            ("accounts_email",    "TEXT"),
            ("bh_name",           "TEXT"),
            ("bh_email",          "TEXT"),
            ("cheque_location",   "TEXT DEFAULT 'Customer'"),
            ("created_at",        "TIMESTAMPTZ DEFAULT NOW()"),
            ("updated_at",        "TIMESTAMPTZ"),
            # Phase 2 — multiple-bounce outcomes & case closure
            ("legal_date",        "DATE"),
            ("legal_reference",   "TEXT"),
            ("rtgs_date",         "DATE"),
            ("rtgs_reference",    "TEXT"),
            ("rtgs_amount",       "NUMERIC"),
            ("closed_date",       "DATE"),
            ("closed_by",         "TEXT"),
            ("close_reason",      "TEXT"),
            ("last_reminder_at",  "TIMESTAMPTZ"),
        ]
        for name, ddl in cheque_cols:
            cur.execute(f"ALTER TABLE cheques ADD COLUMN IF NOT EXISTS {name} {ddl}")

        # Duplicate protection — only enforced when both values are present
        cur.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS uq_cheque_acct_no
            ON cheques (account_number, cheque_number)
            WHERE account_number IS NOT NULL AND cheque_number IS NOT NULL
        """)

        # Staff master (uploaded via Excel)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS staff_master (
                sales_name     TEXT PRIMARY KEY,
                sales_email    TEXT,
                location       TEXT,
                plant          TEXT,
                accounts_email TEXT,
                bh_name        TEXT,
                bh_email       TEXT,
                updated_at     TIMESTAMPTZ DEFAULT NOW()
            )
        """)

        # Event / audit trail (deposit lifecycle, re-deposits)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cheque_events (
                id          SERIAL PRIMARY KEY,
                cheque_id   INTEGER REFERENCES cheques(id) ON DELETE CASCADE,
                action      TEXT,
                action_date DATE,
                bank        TEXT,
                reference   TEXT,
                reason      TEXT,
                created_at  TIMESTAMPTZ DEFAULT NOW()
            )
        """)
        cur.execute("ALTER TABLE cheque_events ADD COLUMN IF NOT EXISTS remarks TEXT")
        cur.execute("ALTER TABLE cheque_events ADD COLUMN IF NOT EXISTS done_by TEXT")

        # Users (login + roles)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id            SERIAL PRIMARY KEY,
                email         TEXT UNIQUE NOT NULL,
                name          TEXT,
                password_hash TEXT NOT NULL,
                role          TEXT NOT NULL DEFAULT 'VIEWER',
                active        BOOLEAN NOT NULL DEFAULT TRUE,
                created_at    TIMESTAMPTZ DEFAULT NOW()
            )
        """)

        # Change / edit log (freeze overrides, staff-master edits, user changes)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS change_log (
                id            SERIAL PRIMARY KEY,
                entity_type   TEXT,
                entity_id     TEXT,
                field_changed TEXT,
                old_value     TEXT,
                new_value     TEXT,
                reason        TEXT,
                changed_by    TEXT,
                changed_at    TIMESTAMPTZ DEFAULT NOW()
            )
        """)

        # Seed the first HO Admin from env, only if no users exist yet
        cur.execute("SELECT COUNT(*) FROM users")
        if cur.fetchone()[0] == 0:
            seed_email = (os.environ.get("ADMIN_EMAIL") or "").strip().lower()
            seed_pw    = os.environ.get("ADMIN_PASSWORD") or ""
            if seed_email and seed_pw:
                cur.execute(
                    "INSERT INTO users (email, name, password_hash, role) "
                    "VALUES (%s,%s,%s,'HO_ADMIN')",
                    (seed_email, "HO Admin", generate_password_hash(seed_pw)),
                )
                print(f"[DB] Seeded first HO Admin: {seed_email}")
            else:
                print("[DB] No users yet and ADMIN_EMAIL/ADMIN_PASSWORD not set — "
                      "set them to seed the first admin.")

        conn.commit()
        cur.close()
    finally:
        conn.close()


with app.app_context():
    try:
        init_db()
    except Exception as _e:
        print(f"[DB] init skipped: {_e}")


# ── Claude prompt ─────────────────────────────────────────────────────────────

PROMPT = """You are scanning an Indian bank cheque. Extract these fields and return them as a single valid JSON object.

Fields:
1. bank_name       — The DRAWER's bank. In India this is the bank name/logo printed at the TOP-LEFT of the cheque. IMPORTANT: ignore any rubber stamps near the bottom — those belong to the collecting/receipt bank, NOT the drawer's bank.
2. account_number  — Read ONLY the digits printed in the "A/c No." field. Read them exactly, digit-for-digit. Do NOT pad, do NOT add leading or trailing zeros, do NOT read the MICR band at the very bottom. If you cannot read it confidently, return null.
3. cheque_number   — The cheque number from the MICR line at the very bottom: take ONLY the LEFTMOST block of digits (usually 6), printed between the ⑆ ⑈ MICR symbols. IGNORE every block to its right (city/bank/branch code, account code, transaction code). These are MICR (E-13B) machine-font digits — read them one digit at a time and look closely at each glyph before committing, ESPECIALLY the first digit (on a blurry, tilted or reflective photo a thin vertical "1" is easily confused with "9" or "7"). If the same cheque number is also printed in ordinary type at the top-right corner of the cheque, use that to cross-check. Return null if you cannot read it confidently.
4. date            — The CHEQUE DATE. Read it ONLY from the row of small digit boxes printed under the labels "D D  M M  Y Y Y Y" (the validity/date boxes, near the "Valid for three months from the date of issue" text). Read the eight boxes strictly left-to-right and group them as DD MM YYYY. Output this raw value as DD/MM/YYYY. CRITICAL: do NOT use any other date printed on the cheque — ignore stationery/print/reference codes such as "CAV/2024/UF 09/12/24", batch/print dates, and the validity wording itself. If the date boxes are blank or unreadable, fall back ONLY to a clearly handwritten date written next to the word "Date". Return null if no cheque date can be determined.
5. cheque_date_iso — The SAME date from field 4, normalised to ISO YYYY-MM-DD (e.g. boxes "1 3 0 5 2 0 2 6" → date "13/05/2026" and cheque_date_iso "2026-05-13"). Must be consistent with field 4. Return null if no date.
6. payee           — Name in the "Pay" field.
7. amount_words    — Amount written in words on the "Rupees" line (e.g. "Five Lakhs Only").
8. amount_numbers  — Amount in the numeric box (e.g. "5,00,000").
9. issuer_name     — Account holder / drawer name printed at the bottom of the cheque.

Rules:
- The photo may be rotated or sideways. First mentally orient the cheque the right way up (bank logo top-left, MICR line along the bottom), then read fields by their printed labels — do NOT rely on absolute screen position.
- Return ONLY a raw JSON object, no markdown fences, no explanation.
- Use null for any field you cannot read confidently. Never guess.

Example (illustrative only — read the actual cheque; date boxes "0 5 0 7 2 0 2 5" → 05/07/2025):
{"bank_name":"ICICI Bank","account_number":"015601500005","cheque_number":"000501","date":"05/07/2025","cheque_date_iso":"2025-07-05","payee":"RDC Concrete","amount_words":"Five Lakhs Only","amount_numbers":"5,00,000","issuer_name":"BRIG K S BHOON"}"""


def build_source(img_bytes, media_type):
    b64 = base64.standard_b64encode(img_bytes).decode("utf-8")
    if media_type == "application/pdf":
        return {"type": "document",
                "source": {"type": "base64", "media_type": "application/pdf", "data": b64}}
    return {"type": "image",
            "source": {"type": "base64", "media_type": media_type, "data": b64}}


def read_cheque(img_bytes, media_type="image/jpeg"):
    message = get_client().messages.create(
        model="claude-sonnet-4-6",
        max_tokens=700,
        messages=[
            {
                "role": "user",
                "content": [
                    build_source(img_bytes, media_type),
                    {"type": "text", "text": PROMPT},
                ],
            }
        ],
    )
    raw = message.content[0].text.strip()
    raw = re.sub(r"^```[a-z]*\n?", "", raw, flags=re.MULTILINE)
    raw = re.sub(r"\n?```$", "", raw, flags=re.MULTILINE)
    return json.loads(raw.strip())


# ── Utilities ─────────────────────────────────────────────────────────────────

def parse_amount(s):
    """'5,00,000' or '₹ 5,00,000.50' → float, else None."""
    if not s:
        return None
    try:
        return float(re.sub(r"[^\d.]", "", str(s)))
    except ValueError:
        return None


def parse_iso_date(s):
    """Return a date for a 'YYYY-MM-DD' string, else None."""
    if not s:
        return None
    try:
        return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def file_ext(filename):
    return filename.rsplit(".", 1)[1].lower() if "." in filename else ""


# ── Auth: helpers, decorators, audit ─────────────────────────────────────────

def current_user():
    """The logged-in user dict {email, name, role} or None."""
    return session.get("user")


@app.context_processor
def inject_user():
    """Make `user` available in every template."""
    return {"user": current_user()}


@app.template_filter("dmy")
def _dmy(d):
    """Render a date/datetime as DD/MM/YYYY (company format); '—' for empty."""
    if not d:
        return "—"
    try:
        return d.strftime("%d/%m/%Y")
    except AttributeError:
        return str(d)


def _wants_json():
    """True for fetch/XHR/API calls; False for top-level page loads."""
    if request.method != "GET":
        return True
    accept = request.headers.get("Accept", "")
    return "application/json" in accept and "text/html" not in accept


def _deny(msg, code):
    if _wants_json():
        return jsonify({"error": msg}), code
    if code == 401 or not current_user():
        return redirect(url_for("login", next=request.path))
    # Logged in but not allowed on this page — send them to a page they can use
    # (their landing) rather than a dead-end "not authorised" screen.
    return redirect(url_for(landing_for(current_user()["role"])))


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            u = current_user()
            if not u:
                return _deny("auth_required", 401)
            if roles and u["role"] not in roles:
                return _deny("forbidden", 403)
            return f(*args, **kwargs)
        return wrapper
    return decorator


def login_required(f):
    """Any authenticated user (any role)."""
    return role_required(*ROLES)(f)


def log_change(entity_type, entity_id, field, old, new, reason=None):
    """Write one row to change_log, stamped with the current user. Best-effort."""
    u = current_user() or {}
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            """INSERT INTO change_log
                 (entity_type, entity_id, field_changed, old_value, new_value, reason, changed_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
            (entity_type, str(entity_id), field,
             None if old is None else str(old),
             None if new is None else str(new),
             reason, u.get("email")),
        )
        conn.commit()
        cur.close()
    finally:
        conn.close()


# ── Routes: auth ──────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        if current_user():
            return redirect(url_for(landing_for(current_user()["role"])))
        return render_template("login.html")

    email = (request.form.get("email") or "").strip().lower()
    password = request.form.get("password") or ""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT email, name, password_hash, role, active FROM users WHERE email=%s",
            (email,),
        )
        row = cur.fetchone()
        cur.close()
    finally:
        conn.close()

    if not row or not row[4] or not check_password_hash(row[2], password):
        return render_template("login.html", error="Invalid email or password."), 401

    session["user"] = {"email": row[0], "name": row[1], "role": row[3]}
    nxt = request.args.get("next") or url_for(landing_for(row[3]))
    return redirect(nxt)


@app.route("/logout", methods=["GET", "POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── Routes: user management (HO Admin only) ───────────────────────────────────

USER_COLS = ["id", "email", "name", "role", "active", "created_at"]


@app.route("/users", methods=["GET"])
@role_required("HO_ADMIN")
def users_page():
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT {', '.join(USER_COLS)} FROM users ORDER BY role, email")
        rows = [dict(zip(USER_COLS, r)) for r in cur.fetchall()]
        cur.close()
    finally:
        conn.close()
    return render_template("users.html", users=rows, roles=ROLES)


@app.route("/users", methods=["POST"])
@role_required("HO_ADMIN")
def users_create():
    d = request.get_json(silent=True) or {}
    email = (d.get("email") or "").strip().lower()
    name  = (d.get("name") or "").strip() or None
    role  = (d.get("role") or "").strip().upper()
    pw    = d.get("password") or ""
    if not email or "@" not in email:
        return jsonify({"error": "A valid email is required"}), 400
    if role not in ROLES:
        return jsonify({"error": f"role must be one of {ROLES}"}), 400
    if len(pw) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO users (email, name, password_hash, role) VALUES (%s,%s,%s,%s) RETURNING id",
            (email, name, generate_password_hash(pw), role),
        )
        new_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        log_change("USER", new_id, "create", None, f"{email} ({role})")
        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        conn.rollback()
        if "users_email_key" in str(e):
            return jsonify({"error": "A user with that email already exists"}), 409
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/users/<int:uid>", methods=["POST"])
@role_required("HO_ADMIN")
def users_update(uid):
    """Toggle active, change role, or reset password."""
    d = request.get_json(silent=True) or {}
    me = current_user()
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT email, role, active FROM users WHERE id=%s", (uid,))
        row = cur.fetchone()
        if not row:
            cur.close()
            return jsonify({"error": "User not found"}), 404
        email, old_role, old_active = row

        if "active" in d:
            new_active = bool(d["active"])
            if email == me["email"] and not new_active:
                cur.close()
                return jsonify({"error": "You cannot deactivate your own account"}), 400
            cur.execute("UPDATE users SET active=%s WHERE id=%s", (new_active, uid))
            log_change("USER", uid, "active", old_active, new_active)
        if "role" in d:
            new_role = str(d["role"]).upper()
            if new_role not in ROLES:
                cur.close()
                return jsonify({"error": f"role must be one of {ROLES}"}), 400
            cur.execute("UPDATE users SET role=%s WHERE id=%s", (new_role, uid))
            log_change("USER", uid, "role", old_role, new_role)
        if d.get("password"):
            if len(d["password"]) < 6:
                cur.close()
                return jsonify({"error": "Password must be at least 6 characters"}), 400
            cur.execute("UPDATE users SET password_hash=%s WHERE id=%s",
                        (generate_password_hash(d["password"]), uid))
            log_change("USER", uid, "password", "***", "***")
        conn.commit()
        cur.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ── Routes: scan UI ──────────────────────────────────────────────────────────

@app.route("/")
@role_required(*UPLOAD_ROLES)
def index():
    return render_template("index.html")


@app.route("/predict", methods=["POST"])
@role_required(*UPLOAD_ROLES)
def predict():
    if "image" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["image"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400
    ext = file_ext(file.filename)
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error": "Unsupported file type"}), 400

    img_bytes = file.read()
    if ext == "pdf":
        media_type = "application/pdf"
    elif ext in {"jpg", "jpeg"}:
        media_type = "image/jpeg"
    else:
        media_type = f"image/{ext}"

    try:
        data = read_cheque(img_bytes, media_type)
        return jsonify({"cheque": data})
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Could not parse cheque data: {e}"}), 500
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# ── Routes: staff master ─────────────────────────────────────────────────────

STAFF_FIELDS = ["sales_name", "sales_email", "location", "plant",
                "accounts_email", "bh_name", "bh_email"]


@app.route("/staff")
@login_required
def staff_list():
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT {', '.join(STAFF_FIELDS)} FROM staff_master ORDER BY sales_name")
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()
    return jsonify([dict(zip(STAFF_FIELDS, r)) for r in rows])


@app.route("/staff/upload", methods=["GET", "POST"])
@role_required("HO_ADMIN")
def staff_upload():
    if request.method == "GET":
        conn = get_db()
        try:
            cur = conn.cursor()
            cur.execute(f"SELECT {', '.join(STAFF_FIELDS)} FROM staff_master ORDER BY sales_name")
            staff = [dict(zip(STAFF_FIELDS, r)) for r in cur.fetchall()]
            cur.close()
        finally:
            conn.close()
        return render_template("staff.html", count=len(staff), staff=staff, fields=STAFF_FIELDS)

    # POST — parse uploaded xlsx
    if "file" not in request.files or request.files["file"].filename == "":
        return jsonify({"error": "No file selected"}), 400
    file = request.files["file"]
    if file_ext(file.filename) not in STAFF_EXTENSIONS:
        return jsonify({"error": "Please upload an .xlsx file"}), 400

    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return jsonify({"error": f"Could not read Excel: {e}"}), 400

    if not rows or len(rows) < 2:
        return jsonify({"error": "Sheet is empty or has no data rows"}), 400

    # Map header → column index (case-insensitive, by expected names)
    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    try:
        idx = {f: header.index(f) for f in STAFF_FIELDS}
    except ValueError:
        return jsonify({
            "error": "Header row must contain columns: " + ", ".join(STAFF_FIELDS)
        }), 400

    conn = get_db()
    saved = 0
    try:
        cur = conn.cursor()
        for r in rows[1:]:
            name = r[idx["sales_name"]] if idx["sales_name"] < len(r) else None
            if not name or str(name).strip() == "":
                continue
            vals = [str(r[idx[f]]).strip() if idx[f] < len(r) and r[idx[f]] is not None else None
                    for f in STAFF_FIELDS]
            cur.execute(
                """
                INSERT INTO staff_master
                    (sales_name, sales_email, location, plant, accounts_email, bh_name, bh_email, updated_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s, NOW())
                ON CONFLICT (sales_name) DO UPDATE SET
                    sales_email=EXCLUDED.sales_email, location=EXCLUDED.location,
                    plant=EXCLUDED.plant, accounts_email=EXCLUDED.accounts_email,
                    bh_name=EXCLUDED.bh_name, bh_email=EXCLUDED.bh_email, updated_at=NOW()
                """,
                vals,
            )
            saved += 1
        conn.commit()
        cur.close()
        return jsonify({"success": True, "saved": saved})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


STAFF_EDITABLE = ["sales_email", "location", "plant", "accounts_email", "bh_name", "bh_email"]


@app.route("/staff/update", methods=["POST"])
@role_required("HO_ADMIN")
def staff_update():
    """Edit one staff record's fields (keyed by sales_name); log each change."""
    d = request.get_json(silent=True) or {}
    name = (d.get("sales_name") or "").strip()
    if not name:
        return jsonify({"error": "sales_name is required"}), 400
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT {', '.join(STAFF_EDITABLE)} FROM staff_master WHERE sales_name=%s", (name,))
        row = cur.fetchone()
        if not row:
            cur.close()
            return jsonify({"error": "Staff not found"}), 404
        old = dict(zip(STAFF_EDITABLE, row))
        sets, vals, changes = [], [], []
        for f in STAFF_EDITABLE:
            if f in d:
                new = (str(d[f]).strip() or None) if d[f] is not None else None
                if new != old[f]:
                    sets.append(f"{f}=%s"); vals.append(new)
                    changes.append((f, old[f], new))
        if not sets:
            cur.close()
            return jsonify({"success": True, "unchanged": True})
        vals.append(name)
        cur.execute(f"UPDATE staff_master SET {', '.join(sets)}, updated_at=NOW() WHERE sales_name=%s", vals)
        conn.commit()
        cur.close()
        for f, o, n in changes:
            log_change("STAFF", name, f, o, n)
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/staff/delete", methods=["POST"])
@role_required("HO_ADMIN")
def staff_delete():
    d = request.get_json(silent=True) or {}
    name = (d.get("sales_name") or "").strip()
    if not name:
        return jsonify({"error": "sales_name is required"}), 400
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM staff_master WHERE sales_name=%s", (name,))
        deleted = cur.rowcount
        conn.commit()
        cur.close()
        if deleted:
            log_change("STAFF", name, "delete", name, None)
        return jsonify({"success": bool(deleted)})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ── Routes: save cheque ──────────────────────────────────────────────────────

@app.route("/accept", methods=["POST"])
@role_required(*UPLOAD_ROLES)
def accept():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "No data received"}), 400

    account_number = (data.get("account_number") or "").strip() or None
    cheque_number  = (data.get("cheque_number") or "").strip() or None

    conn = get_db()
    try:
        cur = conn.cursor()

        # Explicit duplicate check for a friendly message
        if account_number and cheque_number:
            cur.execute(
                "SELECT id FROM cheques WHERE account_number=%s AND cheque_number=%s",
                (account_number, cheque_number),
            )
            existing = cur.fetchone()
            if existing:
                cur.close()
                return jsonify({
                    "error": "duplicate",
                    "message": f"Cheque #{cheque_number} on account {account_number} "
                               f"is already recorded (entry #{existing[0]}).",
                }), 409

        cheque_date_iso  = parse_iso_date(data.get("cheque_date_iso"))
        deposit_due_date = parse_iso_date(data.get("deposit_due_date")) or cheque_date_iso
        amount_value     = parse_amount(data.get("amount_numbers"))

        cur.execute(
            """
            INSERT INTO cheques
                (bank_name, account_number, cheque_number, cheque_date, cheque_date_iso,
                 deposit_due_date, payee, amount_words, amount_numbers, amount_value,
                 issuer_name, status, sales_name, sales_email, location, plant,
                 accounts_email, bh_name, bh_email, cheque_location, updated_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'PENDING',%s,%s,%s,%s,%s,%s,%s,%s, NOW())
            RETURNING id
            """,
            (
                data.get("bank_name"), account_number, cheque_number,
                data.get("date"), cheque_date_iso, deposit_due_date,
                data.get("payee"), data.get("amount_words"), data.get("amount_numbers"),
                amount_value, data.get("issuer_name"),
                data.get("sales_name"), data.get("sales_email"), data.get("location"),
                data.get("plant"), data.get("accounts_email"), data.get("bh_name"),
                data.get("bh_email"), data.get("cheque_location") or "Customer",
            ),
        )
        new_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        conn.rollback()
        # Unique-index violation fallback
        if "uq_cheque_acct_no" in str(e):
            return jsonify({"error": "duplicate",
                            "message": "This cheque is already recorded."}), 409
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ── Routes: report + lifecycle actions ───────────────────────────────────────

REPORT_COLS = [
    "id", "bank_name", "account_number", "cheque_number", "payee",
    "amount_numbers", "amount_value", "cheque_date_iso", "deposit_due_date",
    "status", "sales_name", "location", "plant", "bh_name", "cheque_location",
]

CHEQUE_LOCATIONS = ["Customer", "RDC Accounts", "RDC Sales"]


@app.route("/cheque/<int:cid>/location", methods=["POST"])
@role_required("HO_ADMIN", "ACCOUNTS")
def set_location(cid):
    d = request.get_json(silent=True) or {}
    loc = (d.get("cheque_location") or "").strip()
    if loc not in CHEQUE_LOCATIONS:
        return jsonify({"error": f"cheque_location must be one of {CHEQUE_LOCATIONS}"}), 400
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("UPDATE cheques SET cheque_location=%s, updated_at=NOW() WHERE id=%s", (loc, cid))
        if cur.rowcount == 0:
            cur.close()
            return jsonify({"error": "Cheque not found"}), 404
        conn.commit()
        cur.close()
        return jsonify({"success": True, "cheque_location": loc})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


def _fetch_pending():
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT {', '.join(REPORT_COLS)},
                   (SELECT COUNT(*) FROM cheque_events e
                    WHERE e.cheque_id = cheques.id AND e.action = 'BOUNCED') AS bounce_count
            FROM cheques
            WHERE status IN ('PENDING','BOUNCED','LEGAL')
            ORDER BY deposit_due_date NULLS LAST, id
            """
        )
        cols = REPORT_COLS + ["bounce_count"]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        cur.close()
    finally:
        conn.close()
    return rows


@app.route("/report")
@login_required
def report():
    rows = _fetch_pending()
    today = date.today()
    soon = today + timedelta(days=2)

    overdue, due_soon, upcoming, undated = [], [], [], []
    for r in rows:
        due = r["deposit_due_date"]
        # stale: Indian cheque invalid 3 months after cheque date
        cd = r["cheque_date_iso"]
        r["stale"] = bool(cd and (today - cd).days > 90)
        if due is None:
            undated.append(r)
        elif due < today:
            overdue.append(r)
        elif due <= soon:
            due_soon.append(r)
        else:
            upcoming.append(r)

    def total(group):
        return sum(x["amount_value"] or 0 for x in group)

    groups = [
        ("Crossed deposit date (Overdue)", "overdue", overdue),
        ("Due within 2 days", "soon", due_soon),
        ("Upcoming", "upcoming", upcoming),
        ("No due date set", "undated", undated),
    ]
    return render_template("report.html", groups=groups, total=total,
                           today=today.isoformat(), today_dmy=today.strftime("%d/%m/%Y"),
                           locations=CHEQUE_LOCATIONS)


# ── Routes: dashboard (all cheques, history + filters) ───────────────────────

DASH_COLS = [
    "id", "bank_name", "account_number", "cheque_number", "payee", "issuer_name",
    "amount_numbers", "amount_value", "cheque_date_iso", "deposit_due_date",
    "status", "sales_name", "location", "plant", "bh_name", "cheque_location",
    "deposited_date", "cleared_date",
]

# Filter token -> human label (order shown in the UI)
DASH_FILTERS = [
    ("all",        "All"),
    ("overdue",    "Overdue"),
    ("pending",    "Pending"),
    ("deposited",  "Deposited"),
    ("bounced",    "Bounced"),
    ("redeposited","Re-deposited"),
    ("cleared",    "Cleared"),
    ("legal",      "Legal"),
    ("rtgs",       "RTGS-settled"),
    ("closed",     "Closed"),
]

ACTIVE_STATUSES = ("PENDING", "BOUNCED", "LEGAL", "DEPOSITED")


@app.route("/dashboard")
@login_required
def dashboard():
    today = date.today()
    f_status = (request.args.get("status") or "all").lower()
    q        = (request.args.get("q") or "").strip()
    f_sales  = (request.args.get("sales") or "").strip()
    f_loc    = (request.args.get("location") or "").strip()
    f_plant  = (request.args.get("plant") or "").strip()
    f_from   = parse_iso_date(request.args.get("from"))
    f_to     = parse_iso_date(request.args.get("to"))

    where, params = [], []

    # A cheque is "re-deposited" if its status is DEPOSITED and it has a REDEPOSITED event.
    REDEP_EXISTS = ("EXISTS (SELECT 1 FROM cheque_events e "
                    "WHERE e.cheque_id = cheques.id AND e.action = 'REDEPOSITED')")

    # status / derived filters
    status_map = {"pending": "PENDING", "bounced": "BOUNCED",
                  "cleared": "CLEARED", "legal": "LEGAL", "rtgs": "RTGS_SETTLED",
                  "closed": "CLOSED"}
    if f_status in status_map:
        where.append("status = %s")
        params.append(status_map[f_status])
    elif f_status == "deposited":
        # first-time deposits only — re-deposited ones live in their own bucket
        where.append(f"status = 'DEPOSITED' AND NOT {REDEP_EXISTS}")
    elif f_status == "redeposited":
        where.append(f"status = 'DEPOSITED' AND {REDEP_EXISTS}")
    elif f_status == "overdue":
        where.append("status IN ('PENDING','BOUNCED','LEGAL') AND deposit_due_date < %s")
        params.append(today)

    if q:
        where.append("(payee ILIKE %s OR account_number ILIKE %s OR cheque_number ILIKE %s "
                     "OR issuer_name ILIKE %s)")
        params += [f"%{q}%"] * 4
    if f_sales:
        where.append("sales_name = %s"); params.append(f_sales)
    if f_loc:
        where.append("location = %s"); params.append(f_loc)
    if f_plant:
        where.append("plant = %s"); params.append(f_plant)
    if f_from:
        where.append("COALESCE(deposit_due_date, cheque_date_iso) >= %s"); params.append(f_from)
    if f_to:
        where.append("COALESCE(deposit_due_date, cheque_date_iso) <= %s"); params.append(f_to)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT {', '.join(DASH_COLS)},
                   (SELECT COUNT(*) FROM cheque_events e
                    WHERE e.cheque_id = cheques.id AND e.action = 'BOUNCED') AS bounce_count,
                   {REDEP_EXISTS} AS is_redeposited
            FROM cheques
            {where_sql}
            ORDER BY COALESCE(deposit_due_date, cheque_date_iso) DESC NULLS LAST, id DESC
            """,
            params,
        )
        cols = DASH_COLS + ["bounce_count", "is_redeposited"]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]

        # Summary across ALL cheques (unfiltered), by status
        cur.execute("SELECT status, COUNT(*), COALESCE(SUM(amount_value),0) "
                    "FROM cheques GROUP BY status")
        summary = {s: {"count": c, "amount": float(a)} for s, c, a in cur.fetchall()}
        # Real totals (only true statuses — before adding derived overlays)
        all_count = sum(v["count"] for v in summary.values())
        all_amount = sum(v["amount"] for v in summary.values())

        # Split DEPOSITED into first-time deposits vs re-deposited (status DEPOSITED + event)
        cur.execute(f"SELECT COUNT(*), COALESCE(SUM(amount_value),0) FROM cheques "
                    f"WHERE status='DEPOSITED' AND {REDEP_EXISTS}")
        rc, ra = cur.fetchone()
        summary["REDEP"] = {"count": rc, "amount": float(ra)}
        if "DEPOSITED" in summary:
            summary["DEPOSITED"] = {"count": summary["DEPOSITED"]["count"] - rc,
                                    "amount": summary["DEPOSITED"]["amount"] - float(ra)}
        # Derived overdue count
        cur.execute("SELECT COUNT(*), COALESCE(SUM(amount_value),0) FROM cheques "
                    "WHERE status IN ('PENDING','BOUNCED','LEGAL') AND deposit_due_date < %s",
                    (today,))
        oc, oa = cur.fetchone()
        summary["OVERDUE"] = {"count": oc, "amount": float(oa)}

        # Filter dropdown options
        cur.execute("SELECT DISTINCT sales_name FROM cheques WHERE sales_name IS NOT NULL ORDER BY 1")
        sales_opts = [r[0] for r in cur.fetchall()]
        cur.execute("SELECT DISTINCT location FROM cheques WHERE location IS NOT NULL ORDER BY 1")
        loc_opts = [r[0] for r in cur.fetchall()]
        cur.execute("SELECT DISTINCT plant FROM cheques WHERE plant IS NOT NULL ORDER BY 1")
        plant_opts = [r[0] for r in cur.fetchall()]
        cur.close()
    finally:
        conn.close()

    for r in rows:
        cd = r["cheque_date_iso"]
        # Cheque expiry = 90 days from the cheque date (Indian cheque validity)
        r["expiry_date"] = (cd + timedelta(days=90)) if cd else None
        r["stale"] = bool(cd and (today - cd).days > 90 and r["status"] in ACTIVE_STATUSES)

    total_amount = sum(r["amount_value"] or 0 for r in rows)
    return render_template(
        "dashboard.html", rows=rows, summary=summary, filters=DASH_FILTERS,
        all_count=all_count, all_amount=all_amount,
        active=f_status, q=q, sales_opts=sales_opts, loc_opts=loc_opts, plant_opts=plant_opts,
        sel_sales=f_sales, sel_loc=f_loc, sel_plant=f_plant,
        f_from=request.args.get("from") or "", f_to=request.args.get("to") or "",
        total_amount=total_amount, today=today.isoformat(), locations=CHEQUE_LOCATIONS,
    )


@app.route("/cheque/<int:cid>/deposit", methods=["POST"])
@role_required("HO_ADMIN", "ACCOUNTS")
def mark_deposit(cid):
    d = request.get_json(silent=True) or {}
    dep_date = parse_iso_date(d.get("date")) or date.today()
    bank = (d.get("bank") or "").strip() or None
    ref  = (d.get("reference") or "").strip() or None
    rem  = (d.get("remarks") or "").strip() or None
    return _transition(cid, "DEPOSITED", action="DEPOSITED",
                       action_date=dep_date, bank=bank, reference=ref, remarks=rem,
                       extra_sql="deposited_date=%s, deposit_bank=%s, deposit_reference=%s",
                       extra_vals=(dep_date, bank, ref))


@app.route("/cheque/<int:cid>/redeposit", methods=["POST"])
@role_required("HO_ADMIN", "ACCOUNTS")
def mark_redeposit(cid):
    d = request.get_json(silent=True) or {}
    dep_date = parse_iso_date(d.get("date")) or date.today()
    bank = (d.get("bank") or "").strip() or None
    ref  = (d.get("reference") or "").strip() or None
    rem  = (d.get("remarks") or "").strip() or None
    return _transition(cid, "DEPOSITED", action="REDEPOSITED",
                       action_date=dep_date, bank=bank, reference=ref, remarks=rem,
                       extra_sql="deposited_date=%s, deposit_bank=%s, deposit_reference=%s, "
                                 "bounce_date=NULL, bounce_reason=NULL",
                       extra_vals=(dep_date, bank, ref))


@app.route("/cheque/<int:cid>/clear", methods=["POST"])
@role_required("HO_ADMIN", "ACCOUNTS")
def mark_clear(cid):
    d = request.get_json(silent=True) or {}
    cdate = parse_iso_date(d.get("date")) or date.today()
    rem   = (d.get("remarks") or "").strip() or None
    return _transition(cid, "CLEARED", action="CLEARED", action_date=cdate, remarks=rem,
                       extra_sql="cleared_date=%s", extra_vals=(cdate,))


@app.route("/cheque/<int:cid>/bounce", methods=["POST"])
@role_required("HO_ADMIN", "ACCOUNTS")
def mark_bounce(cid):
    d = request.get_json(silent=True) or {}
    bdate = parse_iso_date(d.get("date")) or date.today()
    reason = (d.get("reason") or "").strip() or None
    rem   = (d.get("remarks") or "").strip() or None
    return _transition(cid, "BOUNCED", action="BOUNCED", action_date=bdate, reason=reason, remarks=rem,
                       extra_sql="bounce_date=%s, bounce_reason=%s", extra_vals=(bdate, reason))


@app.route("/cheque/<int:cid>/legal", methods=["POST"])
@role_required("HO_ADMIN", "ACCOUNTS")
def mark_legal(cid):
    d = request.get_json(silent=True) or {}
    ldate = parse_iso_date(d.get("date")) or date.today()
    ref   = (d.get("reference") or "").strip() or None   # case / FIR number
    rem   = (d.get("remarks") or "").strip() or None
    return _transition(cid, "LEGAL", action="LEGAL", action_date=ldate,
                       reference=ref, remarks=rem,
                       extra_sql="legal_date=%s, legal_reference=%s",
                       extra_vals=(ldate, ref))


@app.route("/cheque/<int:cid>/rtgs", methods=["POST"])
@role_required("HO_ADMIN", "ACCOUNTS")
def mark_rtgs(cid):
    """RTGS / direct payment received — money recovered another way; case closes."""
    d = request.get_json(silent=True) or {}
    rdate  = parse_iso_date(d.get("date")) or date.today()
    ref    = (d.get("reference") or "").strip() or None
    amount = parse_amount(d.get("amount"))
    rem    = (d.get("remarks") or "").strip() or None
    return _transition(cid, "RTGS_SETTLED", action="RTGS_SETTLED", action_date=rdate,
                       reference=ref, remarks=rem,
                       extra_sql="rtgs_date=%s, rtgs_reference=%s, rtgs_amount=%s",
                       extra_vals=(rdate, ref, amount))


@app.route("/cheque/<int:cid>/close", methods=["POST"])
@role_required("HO_ADMIN", "ACCOUNTS")
def mark_close(cid):
    """Explicitly close a case (e.g. a legal case concluded) — stops reminders."""
    d = request.get_json(silent=True) or {}
    cdate  = parse_iso_date(d.get("date")) or date.today()
    reason = (d.get("reason") or "").strip() or None
    rem    = (d.get("remarks") or "").strip() or None
    who    = (current_user() or {}).get("email")
    return _transition(cid, "CLOSED", action="CLOSED", action_date=cdate,
                       reason=reason, remarks=rem,
                       extra_sql="closed_date=%s, closed_by=%s, close_reason=%s",
                       extra_vals=(cdate, who, reason))


@app.route("/cheque/<int:cid>/override", methods=["POST"])
@role_required("HO_ADMIN")
def override_cheque(cid):
    """HO-Admin-only edit of the frozen Cheque Date and/or Amount, with mandatory reason
    and a full before/after audit trail."""
    d = request.get_json(silent=True) or {}
    reason = (d.get("reason") or "").strip()
    if not reason:
        return jsonify({"error": "A reason is required to override locked fields."}), 400

    new_date_raw = d.get("cheque_date")
    new_iso      = parse_iso_date(d.get("cheque_date_iso"))
    new_amt_txt  = d.get("amount_numbers")
    has_amt      = new_amt_txt is not None and str(new_amt_txt).strip() != ""
    has_date     = (new_date_raw is not None and str(new_date_raw).strip() != "") or new_iso is not None
    if not has_amt and not has_date:
        return jsonify({"error": "Nothing to change."}), 400

    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT cheque_date, cheque_date_iso, amount_numbers, amount_value FROM cheques WHERE id=%s",
            (cid,),
        )
        row = cur.fetchone()
        if not row:
            cur.close()
            return jsonify({"error": "Cheque not found"}), 404
        old_date, old_iso, old_amt_txt, old_amt_val = row

        # Only update & log a field if it actually changed.
        sets, vals, changes = [], [], []
        if has_date and new_iso != old_iso:
            new_raw = (str(new_date_raw).strip() if new_date_raw else None) or (
                new_iso.isoformat() if new_iso else None)
            sets += ["cheque_date=%s", "cheque_date_iso=%s"]
            vals += [new_raw, new_iso]
            changes.append(("cheque_date", old_date, new_raw))
        if has_amt:
            new_val = parse_amount(new_amt_txt)
            old_val_f = float(old_amt_val) if old_amt_val is not None else None
            if new_val != old_val_f:
                sets += ["amount_numbers=%s", "amount_value=%s"]
                vals += [str(new_amt_txt).strip(), new_val]
                changes.append(("amount_numbers", old_amt_txt, str(new_amt_txt).strip()))

        if not sets:
            cur.close()
            return jsonify({"success": True, "unchanged": True})

        vals.append(cid)
        cur.execute(f"UPDATE cheques SET {', '.join(sets)}, updated_at=NOW() WHERE id=%s", vals)
        # Event entry on the cheque's own timeline
        cur.execute(
            """INSERT INTO cheque_events (cheque_id, action, action_date, reason, remarks, done_by)
               VALUES (%s,'OVERRIDE',%s,%s,%s,%s)""",
            (cid, date.today(), reason, (d.get("remarks") or "").strip() or None,
             (current_user() or {}).get("email")),
        )
        conn.commit()
        cur.close()
        # Audit log: one row per changed field
        for field, old, new in changes:
            log_change("CHEQUE", cid, field, old, new, reason)
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/cheque/<int:cid>/history")
@login_required
def cheque_history(cid):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            """SELECT action,
                      TO_CHAR(action_date,'DD/MM/YYYY'),
                      bank, reference, reason, remarks, done_by,
                      TO_CHAR(created_at AT TIME ZONE 'Asia/Kolkata','DD/MM/YYYY HH24:MI')
               FROM cheque_events WHERE cheque_id=%s ORDER BY id""",
            (cid,),
        )
        cols = ["action", "action_date", "bank", "reference", "reason", "remarks", "done_by", "logged_at"]
        events = [dict(zip(cols, r)) for r in cur.fetchall()]

        # Edit log (field-level changes to date/amount, with the user who made them)
        cur.execute(
            """SELECT field_changed, old_value, new_value, reason, changed_by,
                      TO_CHAR(changed_at AT TIME ZONE 'Asia/Kolkata','DD/MM/YYYY HH24:MI')
               FROM change_log WHERE entity_type='CHEQUE' AND entity_id=%s ORDER BY id""",
            (str(cid),),
        )
        ecols = ["field", "old", "new", "reason", "changed_by", "changed_at"]
        edits = [dict(zip(ecols, r)) for r in cur.fetchall()]
        cur.close()
    finally:
        conn.close()
    return jsonify({"events": events, "edits": edits})


def _transition(cid, new_status, action, action_date,
                bank=None, reference=None, reason=None, remarks=None,
                extra_sql="", extra_vals=()):
    conn = get_db()
    try:
        cur = conn.cursor()
        set_clause = "status=%s, updated_at=NOW()"
        vals = [new_status]
        if extra_sql:
            set_clause += ", " + extra_sql
            vals.extend(extra_vals)
        vals.append(cid)
        cur.execute(f"UPDATE cheques SET {set_clause} WHERE id=%s", vals)
        if cur.rowcount == 0:
            cur.close()
            return jsonify({"error": "Cheque not found"}), 404
        cur.execute(
            """INSERT INTO cheque_events
                 (cheque_id, action, action_date, bank, reference, reason, remarks, done_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
            (cid, action, action_date, bank, reference, reason, remarks,
             (current_user() or {}).get("email")),
        )
        conn.commit()
        cur.close()
        return jsonify({"success": True, "status": new_status})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ── Email reminders (daily digest to Accounts / BH / HO) ─────────────────────

def smtp_config():
    user = (os.environ.get("SMTP_USER") or "").strip()
    return {
        "host": os.environ.get("SMTP_HOST", "smtp.gmail.com"),
        "port": int(os.environ.get("SMTP_PORT", "587")),
        "user": user,
        "password": os.environ.get("SMTP_APP_PASSWORD", ""),
        "from": (os.environ.get("SMTP_FROM") or user).strip(),
    }


def send_email(to_addr, subject, html):
    cfg = smtp_config()
    if not cfg["user"] or not cfg["password"]:
        raise RuntimeError("SMTP not configured (set SMTP_USER and SMTP_APP_PASSWORD).")
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = cfg["from"]
    msg["To"] = to_addr
    msg.attach(MIMEText(html, "html"))
    with smtplib.SMTP(cfg["host"], cfg["port"], timeout=30) as server:
        server.starttls()
        server.login(cfg["user"], cfg["password"])
        server.sendmail(cfg["from"], [to_addr], msg.as_string())


REMINDER_COLS = ["id", "bank_name", "account_number", "cheque_number", "payee", "issuer_name",
                 "amount_value", "deposit_due_date", "status", "sales_name", "plant",
                 "location", "accounts_email", "bh_email"]


def _open_reminder_rows():
    """Cheques that still need attention: overdue/due-soon PENDING, plus all BOUNCED & LEGAL."""
    today = date.today()
    soon = today + timedelta(days=2)
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT {', '.join(REMINDER_COLS)}
            FROM cheques
            WHERE status IN ('PENDING','BOUNCED','LEGAL')
              AND (status IN ('BOUNCED','LEGAL') OR deposit_due_date <= %s)
            ORDER BY deposit_due_date NULLS LAST, id
            """,
            (soon,),
        )
        rows = [dict(zip(REMINDER_COLS, r)) for r in cur.fetchall()]
        cur.close()
    finally:
        conn.close()
    return rows, today


def _category(r, today):
    if r["status"] == "BOUNCED":
        return "Bounced — awaiting action"
    if r["status"] == "LEGAL":
        return "Legal case — open"
    due = r["deposit_due_date"]
    if due and due < today:
        return "Overdue"
    return "Due within 2 days"


def _digest_html(cheques, today, recipient_label):
    cats = ["Overdue", "Due within 2 days", "Bounced — awaiting action", "Legal case — open"]
    by_cat = {c: [] for c in cats}
    for r in cheques:
        by_cat[_category(r, today)].append(r)

    def money(v):
        return "₹ {:,.0f}".format(v) if v is not None else "—"

    parts = [
        f"""<div style="font-family:Segoe UI,Arial,sans-serif;color:#1a1a1a;max-width:720px">
        <h2 style="margin:0 0 4px">PDC Cheque Reminder</h2>
        <p style="margin:0 0 16px;color:#555">For {recipient_label} · {today.strftime('%d/%m/%Y')} ·
        {len(cheques)} cheque(s) need attention</p>"""
    ]
    total = 0
    for cat in cats:
        items = by_cat[cat]
        if not items:
            continue
        color = {"Overdue": "#c0392b", "Due within 2 days": "#b7791f",
                 "Bounced — awaiting action": "#c0392b", "Legal case — open": "#7c3aed"}[cat]
        parts.append(f'<h3 style="margin:18px 0 6px;color:{color};font-size:15px">{cat} ({len(items)})</h3>')
        parts.append('<table style="border-collapse:collapse;width:100%;font-size:13px">'
                     '<tr style="background:#f3f4f6;text-align:left">'
                     '<th style="padding:6px 8px;border:1px solid #e5e7eb">Due</th>'
                     '<th style="padding:6px 8px;border:1px solid #e5e7eb">Bank / Cheque No</th>'
                     '<th style="padding:6px 8px;border:1px solid #e5e7eb">Customer</th>'
                     '<th style="padding:6px 8px;border:1px solid #e5e7eb">Sales / Plant</th>'
                     '<th style="padding:6px 8px;border:1px solid #e5e7eb;text-align:right">Amount</th></tr>')
        for r in items:
            total += r["amount_value"] or 0
            due = r["deposit_due_date"].strftime("%d/%m/%Y") if r["deposit_due_date"] else "—"
            parts.append(
                f'<tr><td style="padding:6px 8px;border:1px solid #e5e7eb">{due}</td>'
                f'<td style="padding:6px 8px;border:1px solid #e5e7eb">{r["bank_name"] or "—"}<br>'
                f'<span style="color:#888">{r["cheque_number"] or ""}</span></td>'
                f'<td style="padding:6px 8px;border:1px solid #e5e7eb">{r["issuer_name"] or r["payee"] or "—"}</td>'
                f'<td style="padding:6px 8px;border:1px solid #e5e7eb">{r["sales_name"] or "—"}'
                f'<br><span style="color:#888">{r["plant"] or ""}</span></td>'
                f'<td style="padding:6px 8px;border:1px solid #e5e7eb;text-align:right">{money(r["amount_value"])}</td></tr>'
            )
        parts.append("</table>")
    parts.append(f'<p style="margin:16px 0 4px;font-weight:600">Total outstanding: {money(total)}</p>')
    parts.append('<p style="color:#888;font-size:12px;margin-top:18px">'
                 'Open <a href="https://pdc.bhoon.org/dashboard">the dashboard</a> to act on these cheques. '
                 'This is an automated reminder.</p></div>')
    return "".join(parts)


def run_reminders(dry=False):
    rows, today = _open_reminder_rows()
    ho = (os.environ.get("HO_REMINDER_EMAIL") or "").strip()

    buckets = {}  # email -> {label, ids:set, cheques:[]}

    def add(email, label, r):
        if not email:
            return
        key = email.strip().lower()
        if not key:
            return
        b = buckets.setdefault(key, {"label": label, "ids": set(), "cheques": []})
        if r["id"] not in b["ids"]:
            b["ids"].add(r["id"])
            b["cheques"].append(r)

    for r in rows:
        add(r["accounts_email"], "Accounts Incharge", r)
        add(r["bh_email"], "Business Head", r)

    if ho:
        b = buckets.setdefault(ho.lower(), {"label": "HO Credit Control", "ids": set(), "cheques": []})
        for r in rows:
            if r["id"] not in b["ids"]:
                b["ids"].add(r["id"])
                b["cheques"].append(r)

    results, sent_ids = [], set()
    for email, b in buckets.items():
        subject = f"PDC Reminder: {len(b['cheques'])} cheque(s) need attention — {today.strftime('%d/%m/%Y')}"
        if dry:
            results.append({"to": email, "label": b["label"], "count": len(b["cheques"]), "sent": False})
            continue
        try:
            send_email(email, subject, _digest_html(b["cheques"], today, b["label"]))
            results.append({"to": email, "label": b["label"], "count": len(b["cheques"]), "sent": True})
            sent_ids |= b["ids"]
        except Exception as e:
            results.append({"to": email, "label": b["label"], "count": len(b["cheques"]),
                            "sent": False, "error": str(e)})

    if not dry and sent_ids:
        conn = get_db()
        try:
            cur = conn.cursor()
            cur.execute("UPDATE cheques SET last_reminder_at=NOW() WHERE id = ANY(%s)", (list(sent_ids),))
            conn.commit()
            cur.close()
        finally:
            conn.close()

    return {"open_cheques": len(rows), "recipients": len(buckets),
            "ho_configured": bool(ho), "dry_run": dry, "results": results}


@app.route("/cron/reminders")
def cron_reminders():
    """Hit daily by a scheduler. Protected by ?key=<CRON_SECRET>."""
    secret = os.environ.get("CRON_SECRET", "")
    if not secret or request.args.get("key") != secret:
        return jsonify({"error": "forbidden"}), 403
    try:
        return jsonify(run_reminders(dry=(request.args.get("dry") == "1")))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/reminders")
@role_required("HO_ADMIN")
def reminders_page():
    cfg = smtp_config()
    return render_template(
        "reminders.html",
        smtp_ready=bool(cfg["user"] and cfg["password"]),
        smtp_from=cfg["from"] or "(not set)",
        ho_email=os.environ.get("HO_REMINDER_EMAIL") or "(not set)",
        cron_ready=bool(os.environ.get("CRON_SECRET")),
    )


@app.route("/reminders/run", methods=["POST"])
@role_required("HO_ADMIN")
def reminders_run():
    dry = bool((request.get_json(silent=True) or {}).get("dry", False))
    try:
        return jsonify(run_reminders(dry=dry))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Routes: Excel export ─────────────────────────────────────────────────────

@app.route("/export")
@login_required
def export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, bank_name, account_number, cheque_number, cheque_date,
                   cheque_date_iso, deposit_due_date, payee, amount_words,
                   amount_numbers, amount_value, issuer_name, status,
                   deposited_date, deposit_bank, deposit_reference,
                   cleared_date, bounce_date, bounce_reason, cheque_location,
                   sales_name, sales_email, location, plant, bh_name,
                   TO_CHAR(scanned_at AT TIME ZONE 'Asia/Kolkata', 'DD/MM/YYYY HH24:MI')
            FROM cheques
            ORDER BY COALESCE(deposit_due_date, cheque_date_iso) DESC NULLS LAST, id DESC
            """
        )
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cheques"

    HEADERS = [
        "#", "Bank", "Account No", "Cheque No", "Cheque Date (raw)",
        "Cheque Date", "Deposit Due", "Pay To", "Amount (Words)",
        "Amount (text)", "Amount (₹)", "Issuer", "Status",
        "Deposited On", "Deposit Bank", "Deposit Ref",
        "Cleared On", "Bounced On", "Bounce Reason", "Cheque Location",
        "Sales Name", "Sales Email", "Location", "Plant", "BH Name", "Scanned At (IST)",
    ]

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="4F46E5")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border
    ws.row_dimensions[1].height = 30

    alt = PatternFill("solid", fgColor="F5F3FF")
    for ri, row in enumerate(rows, 2):
        shade = alt if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val if val is not None else "")
            c.border = border
            c.alignment = Alignment(vertical="center")
            if isinstance(val, (date, datetime)):
                c.number_format = "DD/MM/YYYY"   # company date format
            if shade:
                c.fill = shade

    widths = [6, 16, 16, 11, 14, 12, 12, 20, 26, 14, 13, 18, 11,
              12, 16, 14, 11, 11, 20, 14, 18, 22, 14, 12, 18, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name="cheques.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
